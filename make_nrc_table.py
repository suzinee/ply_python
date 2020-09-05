"""
20200905
"""
#0 프로젝트 초기화
#Excel Read & Write Lib.
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
"""
write_wb = Workbook()
write_ws1 = write_wb.active
write_ws1 = write_wb.create_sheet('TDP_NRC2')
"""

f = open('NRC_TDP.txt', mode='wt', encoding='utf-8')
 
#1 len(Byte) 입력
nByte = input(">> N-Byte = ")
nBit = 8*int(nByte) 

#2 NRC AND 조건 입력 Byte[n];Bit[n];Value
tc_NRC = input(">> Test Condition (Byte[n];Bit[n];Value) : ")
list_NRC = []
list_NRC = tc_NRC.split(';')
print('len(list_NRC): %s' %len(list_NRC))


#3 nByte*8 경우의 수 생성
for i in range(0,pow(2,nBit)):
    n = i+1
    hx_i = hex(int(i))
    hx_i = hx_i[2:]
    bn_i = bin(int(i))
    bn_i = bn_i[2:].zfill(nBit)
    print ('HEX %s' %hx_i)
    print ('BIN %s' %bn_i)



"""
    for j in range(0,nBit+1):
        k = j
        r = j
        #4-1 NRC 조건에 해당 하는 경우 TC 생성 --> 함수, N-Byte 마다 반복
        if( k % 6 == 0):
            k = 0
            if( r % (nBit) == 0):
                r = 0
            idx_cond1 = (nBit-1) - (int(list_NRC[k])*8 + int(list_NRC[k+1]))
            idx_cond2 = (nBit-1) - (int(list_NRC[k+3])*8 + int(list_NRC[k+4]))
            if bn_i[idx_cond1] in list_NRC[k+2]:     
                if bn_i[idx_cond2] in list_NRC[k+5]:
                    print('==============================')
                    print('NRC : %s' %bn_i)
                    #write_ws1.cell(n,r).value = bn_i[j]
                    f.write(bn_i)
                    f.write('\r\n')
                    print('==============================')
    
            #4-2 NRC 조건 해당 없음  Continue
"""

#5 TDP 저장
"""
f_name = '_nrc_table_200826_2.xlsx'
f_dir = os.getcwd()
f_save = os.path.join(f_dir,f_name)
write_wb.save(f_save)
write_wb.close()
"""
f.close()
